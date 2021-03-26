import configuration
from openpyxl.styles import PatternFill, colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule,CellIsRule
def stylesheet(ws):
    # faire un wrap sur description et nom tache
    # colonne des listes
    ws.column_dimensions['A'].width = 25
    ws["A1"] = "Liste"
    ws["A1"].font = configuration.font
    # colonne des tâches
    ws.column_dimensions['B'].width = 23
    ws["B1"] = "Nom tâche"
    ws["B1"].font = configuration.font
    # colonne des descriptions
    ws.column_dimensions['C'].width = 35
    ws["C1"] = "Description"
    ws["C1"].font = configuration.font
    # colonne des échéances
    ws.column_dimensions['D'].width = 15
    ws["D1"] = "Echéance"
    ws["D1"].font = configuration.font
    # colonne des étiquettes
    ws.column_dimensions['E'].width = 19
    ws["E1"] = "Etiquettes"
    ws["E1"].font = configuration.font
    # colonne des équipes
    ws["F1"] = "Equipe"
    ws["F1"].font = configuration.font

def wrapcolumn(ws, longueur):
    for row in ws.iter_rows(min_row=2, max_row=longueur+1, min_col=2, max_col=5):
        for cell in row:
            cell.alignment = configuration.wrap
def dateColumn(ws,longueur):
    for row in ws.iter_rows(min_row=2, max_row=longueur+1, min_col=4):
        for cell in row:
            cell.number_format = 'dd-mm-yyyy'

def inserttasks(ws, listecards, listemembres, listelistes):
    # dico d'info à mettre dans l'excel pour une card
    card = {}
    for carte in listecards:
        membreCard = []
        lab = ""
        mem = ""
        # mettre les infos direct de la carte

        # recuperation du nom de la tâche
        card['B'] = carte["name"]

        # récuperation de la description
        card['C'] = carte["desc"]

        # récuperation de la date d'échéance + formatage
        d = configuration.dateparser.parse(str(carte["due"]))
        if d is None:
            card['D'] = ""
        else:
            card['D'] = configuration.datetime.date(d.year,d.month,d.day)

        # récuperation des étiquettes de la tâche + formatage en string
        for label in carte["labels"]:
            if lab == "":
                lab = str(label["name"])
            else:
                lab = str(lab) + "," + str(label["name"])
        card['E'] = lab

        # chercher les id des différentes donnée externe à la carte

        # récupération de la liste de la tâche
        for liste in listelistes:
            if carte["idList"] == liste["id"]:
                card['A'] = liste["name"]

        # récuperation des membres assignés
        for membre in listemembres:
            for id in carte["idMembers"]:
                if id == membre["id"]:
                    membreCard.append(membre["initials"])

        # formatage en string de la liste des membres
        for item in membreCard:
            if mem == "":
                mem = str(item)
            else:
                mem = str(mem) + "," + str(item)
        card['F'] = mem
        ws.append(card)
        card.clear()

def addfilter(ws,longueur):
    ws.auto_filter.ref = "A1:F"+str(longueur)
    ws.auto_filter.add_sort_condition("A1:F"+str(longueur))

def addrules(ws,longueur):
   #rule pour les dates inférieures
   redbackground = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
   diff = DifferentialStyle(fill=redbackground)
   #['AND(NOT(ISEMPTY($D1));($D1<TODAY()))']
   rule5 = Rule(type="expression", dxf=diff, formula=['AND(NOT(ISBLANK($D2)),($D2<TODAY()))'])
   ws.conditional_formatting.add("D2:D466", rule5)
   #rule pour la liste en cours
   ybackground = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
   diff = DifferentialStyle(fill=ybackground)
   rule1 = Rule(type="expression", dxf=diff, formula=['$A2="En cours"'])
   ws.conditional_formatting.add("A2:F"+str(longueur+1), rule1)
   #rule pour la liste bolqué
   obackground = PatternFill(start_color='FF9933', end_color='FF9933', fill_type='solid')
   diff = DifferentialStyle(fill=obackground)
   rule2 = Rule(type="expression", dxf=diff, formula=['$A2="Bloqué"'])
   ws.conditional_formatting.add("A2:F"+str(longueur+1), rule2)
   #rule pour la liste presque fini
   lgbackground = PatternFill(start_color='99FF33', end_color='99FF33', fill_type='solid')
   diff = DifferentialStyle(fill=lgbackground)
   rule3 = Rule(type="expression", dxf=diff, formula=['$A2="Presque fini"'])
   ws.conditional_formatting.add("A2:F"+str(longueur+1), rule3)
   #rule pour la liste terminé
   gbackground = PatternFill(start_color='00CC00', end_color='00CC00', fill_type='solid')
   diff = DifferentialStyle(fill=gbackground)
   rule4 = Rule(type="expression", dxf=diff)
   rule4.formula = ['$A2="Terminé !"']
   ws.conditional_formatting.add("A2:F"+str(longueur+1), rule4)


